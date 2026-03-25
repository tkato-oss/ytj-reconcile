#!/usr/bin/env python3
"""YTJ 請求書突合ツール v1.2 — UI安定版"""

import streamlit as st
import pdfplumber
import openpyxl
import math, io, re, csv, zipfile
from dataclasses import dataclass, field
from typing import Optional

st.set_page_config(page_title="YTJ 請求書突合ツール", page_icon="🔍", layout="wide")
PASSWORD = "ytj2026"
YTJ_RED = "#B71C1C"

st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@300;400;500;600;700&display=swap');
.stApp { background: #FAFAFA; font-family: 'Noto Sans JP', sans-serif; }
#MainMenu {visibility:hidden;} footer {visibility:hidden;} .stDeployButton {display:none;}
</style>""", unsafe_allow_html=True)

# ═══ Data ═══
@dataclass
class WorkReport:
    staff_name:str=""; staff_id:str=""; month:str=""
    teaching:float=0; assist:float=0
    reg_teach:float=0; sub_teach:float=0
    pk:float=0; op:float=0; other:float=0; sub_assist:float=0

@dataclass
class Invoice:
    name:str=""; teaching_h:float=0; teaching_price:int=0; teaching_amt:int=0
    assist_items:list=field(default_factory=list)
    assist_h:float=0; assist_amt:int=0
    subtotal:int=0; tax:int=0; wh_tax:int=0; total:int=0

@dataclass
class Check:
    step:str; name:str; ok:bool; expected:str; actual:str

@dataclass
class Result:
    name:str; sid:str=""; month:str=""; total:int=0
    checks:list=field(default_factory=list)
    work:Optional[WorkReport]=None; inv:Optional[Invoice]=None; error:str=""
    @property
    def fail_count(self): return sum(1 for c in self.checks if not c.ok)

# ═══ Ledger ═══
def parse_ledger(file_bytes, filename):
    ledger = {}
    if filename.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            try:
                raw_id = row[0].value
                if raw_id is None: continue
                sid = str(int(float(raw_id)))
                sei = str(row[1].value or "")
                mei = str(row[2].value or "")
                tp = int(round(float(row[6].value or 0)))
                ap = int(round(float(row[8].value or 0)))
                ledger[sid] = {"name": f"{sei}{mei}", "tp": tp, "ap": ap}
            except: continue
    else:
        text = file_bytes.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        first = True
        for row in reader:
            if not row or len(row) < 9: continue
            if first:
                if "ID" in str(row[0]) or "姓" in str(row[1]): first = False; continue
            try:
                sid = str(int(float(str(row[0]).strip())))
                tp = int(round(float(str(row[6]).replace(",","").strip())))
                ap = int(round(float(str(row[8]).replace(",","").strip())))
                ledger[sid] = {"name": f"{row[1]}{row[2]}", "tp": tp, "ap": ap}
            except: continue
    return ledger

# ═══ ZIP ═══
def extract_pdfs(zip_bytes):
    files = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for n in zf.namelist():
            bn = n.split("/")[-1]
            if bn.startswith(".") or bn.startswith("__") or bn.startswith("~$"): continue
            if n.lower().endswith(".pdf"):
                data = zf.read(n)
                f = io.BytesIO(data); f.name = bn; f.getvalue = lambda d=data: d
                files.append(f)
    return files

def extract_excels(zip_bytes):
    files = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for n in zf.namelist():
            bn = n.split("/")[-1]
            if bn.startswith(".") or bn.startswith("__") or bn.startswith("~$"): continue
            if n.lower().endswith(".xlsx"):
                data = zf.read(n)
                f = io.BytesIO(data); f.name = bn; f.getvalue = lambda d=data: d
                files.append(f)
    return files

# ═══ Excel Reader ═══
def find_sheet(wb, month_str):
    m = re.search(r'(\d+)月', month_str)
    if not m: return None
    mn = m.group(1)
    cands = []
    for sn in wb.sheetnames:
        s = sn.strip()
        if f"{mn}月" in s:
            idx = s.find(f"{mn}月")
            if idx > 0 and s[idx-1].isdigit(): continue
            cands.append(sn)
    if not cands: return None
    for c in cands:
        if "2026" in c: return wb[c]
    return wb[cands[0]]

def read_excel(file_bytes, month):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = find_sheet(wb, month)
    if ws is None: raise ValueError(f"「{month}」シートなし")
    d = WorkReport()
    a1 = str(ws['A1'].value or ""); a2 = str(ws['A2'].value or ""); a3 = str(ws['A3'].value or "")
    for p in ["スタッフ名：","スタッフ名:"]: a1=a1.replace(p,"")
    d.staff_name = a1.strip()
    for p in ["スタッフID：","スタッフID:"]: a2=a2.replace(p,"")
    d.staff_id = a2.strip()
    for p in ["担当月：","担当月:"]: a3=a3.replace(p,"")
    d.month = a3.strip()
    # Fallback: if name/ID is blank, search other sheets
    if not d.staff_name or not d.staff_id:
        for sn in wb.sheetnames:
            try:
                other = wb[sn]
                oa1 = str(other['A1'].value or ""); oa2 = str(other['A2'].value or "")
                for p in ["スタッフ名：","スタッフ名:"]: oa1=oa1.replace(p,"")
                for p in ["スタッフID：","スタッフID:"]: oa2=oa2.replace(p,"")
                if oa1.strip() and not d.staff_name: d.staff_name = oa1.strip()
                if oa2.strip() and not d.staff_id: d.staff_id = oa2.strip()
                if d.staff_name and d.staff_id: break
            except: continue
    sr = None
    for row in ws.iter_rows(min_row=30, max_row=65):
        if row[0].value and "合計勤務時間" in str(row[0].value): sr=row[0].row; break
    if sr is None: raise ValueError("合計勤務時間行なし")
    def sf(v):
        if v is None: return 0.0
        try: return float(v)
        except: return 0.0
    d.reg_teach=sf(ws.cell(row=sr,column=7).value); d.pk=sf(ws.cell(row=sr,column=8).value)
    d.op=sf(ws.cell(row=sr,column=9).value); d.other=sf(ws.cell(row=sr,column=10).value)
    d.sub_teach=sf(ws.cell(row=sr,column=13).value); d.sub_assist=sf(ws.cell(row=sr,column=14).value)
    d.teaching=d.reg_teach+d.sub_teach; d.assist=d.pk+d.op+d.other+d.sub_assist
    return d

# ═══ PDF Reader ═══
def nums(row):
    r=[]
    for c in row:
        if not c: continue
        x=str(c).replace(",","").replace("(","").replace(")","").replace("（","").replace("）","").strip()
        try: r.append(float(x))
        except: pass
    return r

def read_pdf(file_bytes):
    d = Invoice()
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        tables = pdf.pages[0].extract_tables()
    if not tables: raise ValueError("PDF読取失敗")
    TKW=["レッスン担当","指導"]; AKW=["補助","PK補助","レッスン補助","オープン業務","OP清掃","清掃"]
    SKW=["日付","内容","数量","単価","金額","氏名","発行","振込","銀行","支店","種別","口座","請求","INVOICE","住所","電話","メール","インボイス","備考"]
    for row in tables[0]:
        row=[str(c).strip() if c else "" for c in row]; rt=" ".join(row)
        is_t=any(k in rt for k in TKW); is_a=any(k in rt for k in AKW) and not is_t; is_s=any(k in rt for k in SKW)
        if "氏名" in rt:
            for i,c in enumerate(row):
                if "氏名" in c and i+1<len(row) and row[i+1]: d.name=row[i+1].strip(); break
        if is_t and not is_s:
            n=nums(row)
            if len(n)>=3: d.teaching_h=n[0]; d.teaching_price=int(n[1]); d.teaching_amt=int(n[2])
        elif is_a and not is_s:
            nm="補助"
            for k in AKW:
                if k in rt: nm=k; break
            n=nums(row)
            if len(n)>=3: d.assist_items.append({"name":nm,"h":n[0],"p":int(n[1]),"a":int(n[2])})
        if "小計" in rt:
            n=nums(row)
            if n: d.subtotal=int(n[0])
        if "消費税" in rt and "源泉" not in rt:
            n=nums(row)
            if n: d.tax=int(n[0])
        if "源泉" in rt:
            n=nums(row)
            if n: d.wh_tax=int(n[0])
        if "合計" in rt and "小計" not in rt and "勤務" not in rt:
            n=nums(row)
            if n: d.total=int(n[0])
    d.assist_h=sum(a["h"] for a in d.assist_items)
    d.assist_amt=sum(a["a"] for a in d.assist_items)
    return d

# ═══ Checks ═══
def run_checks(w, inv, led=None):
    r=[]
    r.append(Check("Step1","指導時間",abs(w.teaching-inv.teaching_h)<0.01,f"{inv.teaching_h}h",f"{w.teaching}h"))
    r.append(Check("Step1","補助時間",abs(w.assist-inv.assist_h)<0.01,f"{inv.assist_h}h",f"{w.assist}h"))
    if led:
        r.append(Check("Step2","指導単価",led["tp"]==inv.teaching_price,f"¥{led['tp']:,}",f"¥{inv.teaching_price:,}"))
        aok=all(a["p"]==led["ap"] for a in inv.assist_items) if inv.assist_items else True
        r.append(Check("Step2","補助単価",aok,f"¥{led['ap']:,}",f"¥{inv.assist_items[0]['p']:,}" if inv.assist_items else "—"))
    et=inv.teaching_h*inv.teaching_price
    r.append(Check("Step3","指導金額",abs(et-inv.teaching_amt)<1,f"¥{et:,.0f}",f"¥{inv.teaching_amt:,}"))
    for a in inv.assist_items:
        ea=a["h"]*a["p"]
        r.append(Check("Step3",f"{a['name']}金額",abs(ea-a["a"])<1,f"¥{ea:,.0f}",f"¥{a['a']:,}"))
    es=inv.teaching_amt+inv.assist_amt
    r.append(Check("Step3","小計",es==inv.subtotal,f"¥{es:,}",f"¥{inv.subtotal:,}"))
    etx=math.floor(inv.subtotal*0.10)
    r.append(Check("Step4","消費税",etx==inv.tax,f"¥{etx:,}",f"¥{inv.tax:,}"))
    if inv.subtotal>100000:
        ew=math.floor(inv.subtotal*0.1021)
        r.append(Check("Step4","源泉徴収税",ew==inv.wh_tax,f"¥{ew:,}",f"¥{inv.wh_tax:,}"))
    else:
        r.append(Check("Step4","源泉徴収税",inv.wh_tax==0,"¥0",f"¥{inv.wh_tax:,}"))
    eall=inv.subtotal+inv.tax-inv.wh_tax
    r.append(Check("Step5","合計",eall==inv.total,f"¥{eall:,}",f"¥{inv.total:,}"))
    return r

# ═══ Matching ═══
def norm(s): return s.replace("　","").replace(" ","").strip()
def get_id(fn):
    n=re.findall(r'\d{4}',fn)
    return n[0] if n else None

def fuzzy_match(a, b):
    """Check if names are similar enough (handles 萌花 vs 萌華 etc.)"""
    if not a or not b: return False
    if a == b: return True
    if a in b or b in a: return True
    # Check character overlap: if 2+ chars match out of shorter name, consider it a match
    if len(a) >= 2 and len(b) >= 2:
        common = sum(1 for c in a if c in b)
        shorter = min(len(a), len(b))
        if common >= shorter - 1 and common >= 2:  # Allow 1 char difference
            return True
    return False

def get_name_from_wb(file_bytes):
    """Get staff name from any sheet in workbook"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        for sn in wb.sheetnames:
            a1 = str(wb[sn]['A1'].value or "")
            name = norm(a1.replace("スタッフ名：","").replace("スタッフ名:",""))
            if name: return name
    except: pass
    return ""

def match(wfs, pfs):
    pairs=[]; used=set()
    for wf in wfs:
        wid=get_id(wf.name); best=None
        # 1st: ID match
        if wid:
            for pf in pfs:
                if pf.name in used: continue
                if wid in pf.name: best=pf; break
        # 2nd: Name match (search all sheets, fuzzy)
        if best is None:
            wn = get_name_from_wb(wf.getvalue())
            if wn:
                for pf in pfs:
                    if pf.name in used: continue
                    try:
                        inv=read_pdf(pf.getvalue()); pn=norm(inv.name)
                        if fuzzy_match(wn, pn): best=pf; break
                    except: continue
            # 3rd: Filename match
            if best is None and wn:
                for pf in pfs:
                    if pf.name in used: continue
                    pfn = norm(pf.name.replace(".pdf",""))
                    if fuzzy_match(wn, pfn): best=pf; break
        if best: used.add(best.name)
        pairs.append((wf,best))
    return pairs

# ═══════════════════
# UI
# ═══════════════════
if "auth" not in st.session_state: st.session_state.auth=False
if not st.session_state.auth:
    st.markdown("### 🔍 YTJ 請求書突合ツール")
    st.caption("パスワードを入力してください")
    pw=st.text_input("パスワード",type="password")
    if st.button("ログイン",type="primary"):
        if pw==PASSWORD: st.session_state.auth=True; st.rerun()
        else: st.error("パスワードが違います")
    st.stop()

st.markdown("### 🔍 YTJ 請求書突合ツール")
st.caption("Invoice Reconciliation v1.2")

if "results" not in st.session_state: st.session_state.results=None
if "sel" not in st.session_state: st.session_state.sel=None

# ═══ Upload ═══
if st.session_state.results is None:
    st.subheader("請求書の突合チェック")
    st.caption("3つのファイルをアップロードして突合を実行してください")

    st.markdown("**① 稼働報告書（Excel / ZIP）**")
    st.caption("複数ファイルを選択、またはZIPにまとめてアップロードできます")
    wuploads=st.file_uploader("稼働報告書",type=["xlsx","zip"],accept_multiple_files=True,key="w",label_visibility="collapsed")
    wfs=[]
    if wuploads:
        for f in wuploads:
            if f.name.lower().endswith(".zip"):
                wfs.extend(extract_excels(f.getvalue()))
            else:
                wfs.append(f)
        st.success(f"✅ {len(wfs)}件: " + ", ".join(f.name for f in wfs))

    st.markdown(""); st.markdown(""); st.divider()

    st.markdown("**② 請求書（PDF / ZIP）**")
    st.caption("PDFを複数選択、またはZIPにまとめてアップロードできます")
    iuploads=st.file_uploader("請求書",type=["pdf","zip"],accept_multiple_files=True,key="i",label_visibility="collapsed")
    pdfs=[]
    if iuploads:
        for f in iuploads:
            if f.name.lower().endswith(".zip"):
                pdfs.extend(extract_pdfs(f.getvalue()))
            else:
                pdfs.append(f)
        st.success(f"✅ {len(pdfs)}件: " + ", ".join(f.name for f in pdfs))

    st.markdown(""); st.markdown(""); st.divider()

    st.markdown("**③ 従業員台帳（CSV / Excel）**")
    st.caption("スプレッドシートからエクスポートしてください。新規加入・脱退があれば最新版をアップしてください。")
    lf=st.file_uploader("台帳",type=["csv","xlsx"],key="l",label_visibility="collapsed")
    ledger={}
    if lf:
        try:
            ledger=parse_ledger(lf.getvalue(),lf.name)
            st.success(f"✅ 従業員台帳: {len(ledger)}名")
        except Exception as e:
            st.error(f"台帳エラー: {e}")

    month=None
    if wfs: month=st.text_input("📅 対象月（例: 2月）",value="2月")

    if wfs and pdfs:
        st.info(f"📊 稼働報告書:{len(wfs)}件　📄 請求書:{len(pdfs)}件　📋 台帳:{len(ledger)}名")

    if wfs and pdfs and month:
        if st.button("🚀 突合チェックを実行",type="primary",use_container_width=True):
            results=[]; prog=st.progress(0)
            pairs=match(wfs,pdfs)
            for idx,(wf,pf) in enumerate(pairs):
                prog.progress((idx+1)/len(pairs))
                pr=Result(name="",sid="",month="",total=0)
                try:
                    w=read_excel(wf.getvalue(),month)
                    pr.name=w.staff_name;pr.sid=w.staff_id;pr.month=w.month;pr.work=w
                except Exception as e:
                    pr.name=wf.name;pr.error=f"稼働報告書エラー: {e}";results.append(pr);continue
                if pf is None:
                    pr.error="請求書が見つかりません";results.append(pr);continue
                try:
                    inv=read_pdf(pf.getvalue());pr.total=inv.total;pr.inv=inv
                except Exception as e:
                    pr.error=f"請求書エラー: {e}";results.append(pr);continue
                led=ledger.get(str(w.staff_id).strip(),None)
                pr.checks=run_checks(w,inv,led);results.append(pr)
            prog.empty();st.session_state.results=results;st.rerun()

# ═══ Results ═══
elif st.session_state.sel is None:
    results=st.session_state.results
    total=len(results);err=sum(1 for r in results if r.fail_count>0 or r.error);ok=total-err

    c1,c2,c3=st.columns(3)
    c1.metric("対象者数",f"{total}名")
    c2.metric("問題なし",f"{ok}名")
    c3.metric("要確認",f"{err}名")

    st.divider()
    filt=st.radio("表示",["全員","要確認のみ","OKのみ"],horizontal=True,label_visibility="collapsed")
    if filt=="要確認のみ": fl=[r for r in results if r.fail_count>0 or r.error]
    elif filt=="OKのみ": fl=[r for r in results if r.fail_count==0 and not r.error]
    else: fl=results

    st.markdown("")
    # Header
    hc1,hc2,hc3,hc4,hc5=st.columns([0.5,2,1,1.5,1.5])
    hc1.caption("No."); hc2.caption("氏名（ID）"); hc3.caption("対象月"); hc4.caption("請求額"); hc5.caption("判定")
    st.divider()

    for idx,r in enumerate(fl):
        ri=results.index(r)
        status="🔴 ERR" if r.error else ("✅ OK" if r.fail_count==0 else f"⚠️ {r.fail_count}件")
        amount=f"¥{r.total:,}" if r.total else "—"
        
        col1,col2,col3,col4,col5=st.columns([0.5,2,1,1.5,1.5])
        col1.write(f"{idx+1}")
        col2.write(f"**{r.name}**（{r.sid}）")
        col3.write(r.month if r.month else "—")
        col4.write(amount)
        
        if r.error:
            col5.write(f"🔴 {r.error[:20]}...")
        elif r.fail_count==0:
            col5.write("✅ OK")
        else:
            if col5.button(f"⚠️ {r.fail_count}件 → 詳細",key=f"d{ri}"):
                st.session_state.sel=ri;st.rerun()
        
        st.markdown("")  # 行間のスペース

    st.divider()
    if st.button("← 別のファイルで突合する"):
        st.session_state.results=None;st.session_state.sel=None;st.rerun()

# ═══ Detail ═══
else:
    r=st.session_state.results[st.session_state.sel]
    if st.button("← 一覧に戻る"):
        st.session_state.sel=None;st.rerun()

    st.subheader(f"{r.name}（ID: {r.sid}）")
    st.caption(f"{r.month}分　|　請求額 ¥{r.total:,}　|　不一致 {r.fail_count}件")
    st.divider()

    # Header row
    hc1,hc2,hc3,hc4,hc5=st.columns([0.8,2,1.5,1.5,0.5])
    hc1.caption("STEP"); hc2.caption("チェック項目"); hc3.caption("期待値"); hc4.caption("実際値"); hc5.caption("判定")
    st.divider()

    for ch in r.checks:
        icon="✅" if ch.ok else "❌"
        c1,c2,c3,c4,c5=st.columns([0.8,2,1.5,1.5,0.5])
        c1.caption(ch.step)
        c2.write(f"**{ch.name}**" if not ch.ok else ch.name)
        c3.write(ch.expected)
        if ch.ok:
            c4.write(ch.actual)
        else:
            c4.write(f"**:red[{ch.actual}]**")
        c5.write(icon)
        st.markdown("")  # 行間のスペース

    if r.work:
        w=r.work
        st.divider()
        st.markdown("**📊 稼働報告書（履歴）内訳**")
        c1,c2=st.columns(2)
        with c1:
            st.write(f"レギュラー指導 (G列): **{w.reg_teach}h**")
            st.write(f"代講指導 (M列): **{w.sub_teach}h**")
            st.write(f"→ **指導合計: {w.teaching}h**")
        with c2:
            st.write(f"PK補助 (H列): **{w.pk}h**")
            st.write(f"OP清掃 (I列): **{w.op}h**")
            st.write(f"その他 (J列): **{w.other}h**")
            st.write(f"代講補助 (N列): **{w.sub_assist}h**")
            st.write(f"→ **補助合計: {w.assist}h**")

st.divider()
st.caption("YTJ × PORTAMENT")
